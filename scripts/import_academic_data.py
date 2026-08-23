"""
Import du referentiel academique national depuis
mahay_universites_mentions_filieres_recensement.xlsx vers les tables
Universite / Faculte(Composante) / Domaine / Mention / Filiere /
ProgrammeUniversitaire.

A LANCER APRES la migration b8f4d1c6a2e7 (qui cree la table Domaine).

Perimetre (decide avec Jake) : uniquement les 6 universites PUBLIQUES
du fichier. Les 22 etablissements prives sont ignores pour l'instant
(colonne Universite non presente dans PERIMETRE_UNIVERSITES_PUBLIQUES
plus bas -> simplement sautee ligne par ligne).

Comportement DIFFERENT selon que l'universite a deja des Filiere en
base ou non (voir analyse presentee a Jake avant ce script) :

- Universite "vide" (aucune Filiere existante -- Fianarantsoa,
  Mahajanga, Toliara, Antsiranana au moment ou ce script est ecrit) :
  import complet et direct. Composante/Domaine/Mention/Filiere sont
  crees (ou retrouves s'ils existent deja -- idempotent) et relies.

- Universite "curatee a la main" (a deja des Filiere verifiees avec
  sources citees dans les migrations -- Toamasina, Antananarivo au
  moment ou ce script est ecrit) : CE SCRIPT NE CREE JAMAIS DE
  NOUVELLE FILIERE POUR CES UNIVERSITES, et ne renomme/supprime rien.
  Il tente uniquement de rattacher un Domaine aux Mention deja
  existantes (rattachement additif, jamais un remplacement), en
  utilisant les lignes du fichier Excel concernant cette universite
  pour retrouver, par correspondance EXACTE de nom normalise, la
  Filiere existante correspondante. Toute ligne du fichier sans
  correspondance certaine est listee dans le rapport final pour revue
  admin -- jamais devinee (voir §44 du brief refonte academique).

Idempotent : peut etre relance sans creer de doublon (tout passe par
un get_or_create base sur le nom normalise).

Usage :
    python -m scripts.import_academic_data [--fichier CHEMIN.xlsx] [--dry-run]
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from sqlmodel import Session, select

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database import engine  # noqa: E402
from app.models import Domaine, Faculte, Filiere, Mention, ProgrammeUniversitaire, Universite  # noqa: E402

FICHIER_PAR_DEFAUT = "mahay_universites_mentions_filieres_recensement.xlsx"

# Perimetre : uniquement les 6 universites publiques (decide avec Jake
# le 22/08/2026). Compare par nom NORMALISE (voir normaliser() plus
# bas) -- pas besoin de faire correspondre exactement les accents ici,
# ils sont retires avant comparaison.
PERIMETRE_UNIVERSITES_PUBLIQUES = {
    "universite d'antananarivo",
    "universite d'antsiranana",
    "universite de fianarantsoa",
    "universite de mahajanga",
    "universite de toamasina",
    "universite de toliara",
}


def normaliser(texte: str | None) -> str:
    """Normalise un texte pour COMPARAISON uniquement (jamais pour
    l'affichage/le stockage) : accents retires, apostrophes
    typographiques uniformisees, espaces multiples reduits, casse
    ignoree. Permet de faire correspondre 'Université d'Antananarivo'
    (fichier Excel) et 'Universite d'Antananarivo' (base actuelle,
    sans accent), ou 'Physique et applications' / 'Physique et
    Applications' (collision de casse presente dans le fichier)."""
    if not texte:
        return ""
    texte = texte.strip().replace("\u2019", "'")
    texte = unicodedata.normalize("NFKD", texte)
    texte = "".join(c for c in texte if not unicodedata.combining(c))
    texte = re.sub(r"\s+", " ", texte)
    return texte.lower()


@dataclass
class Rapport:
    universites_hors_perimetre: set = field(default_factory=set)
    domaines_crees: list = field(default_factory=list)
    mentions_creees: list = field(default_factory=list)
    mentions_domaine_ambigu: dict = field(default_factory=lambda: defaultdict(set))
    mentions_domaine_rattache: list = field(default_factory=list)
    facultes_creees: list = field(default_factory=list)
    filieres_creees: list = field(default_factory=list)
    programmes_crees: int = 0
    # --- Universites "curatees" (jamais de creation de Filiere) ---
    filieres_existantes_rattachees_domaine: list = field(default_factory=list)
    lignes_sans_correspondance: list = field(default_factory=list)

    def imprimer(self) -> None:
        print("\n=== RAPPORT D'IMPORT ===\n")
        if self.universites_hors_perimetre:
            print(f"Universites hors perimetre (ignorees) : {sorted(self.universites_hors_perimetre)}")
        print(f"Domaines crees : {len(self.domaines_crees)}")
        print(f"Mentions creees : {len(self.mentions_creees)}")
        print(f"Mentions -> Domaine rattache : {len(self.mentions_domaine_rattache)}")
        if self.mentions_domaine_ambigu:
            print(f"Mentions avec Domaine AMBIGU (non rattache, a trancher en admin) : {len(self.mentions_domaine_ambigu)}")
            for nom_mention, domaines in sorted(self.mentions_domaine_ambigu.items()):
                print(f"   - {nom_mention} : {sorted(domaines)}")
        print(f"Composantes (Faculte) creees : {len(self.facultes_creees)}")
        for u, f in self.facultes_creees:
            print(f"   - [{u}] {f}")
        print(f"Filieres creees (universites sans donnees existantes) : {len(self.filieres_creees)}")
        for u, f, fil in self.filieres_creees:
            print(f"   - [{u} / {f}] {fil}")
        print(f"Liens ProgrammeUniversitaire crees : {self.programmes_crees}")
        print(f"\n--- Universites deja curatees a la main (aucune Filiere creee) ---")
        print(f"Filieres existantes rattachees a un Domaine : {len(self.filieres_existantes_rattachees_domaine)}")
        for u, fil, dom in self.filieres_existantes_rattachees_domaine:
            print(f"   - [{u}] {fil} -> domaine {dom}")
        print(f"Lignes Excel SANS correspondance certaine (a revoir dans /admin/referentiel) : {len(self.lignes_sans_correspondance)}")
        for u, comp, dom, ment, parc in self.lignes_sans_correspondance:
            print(f"   - [{u} / {comp}] {dom} > {ment} > {parc}")


def lire_lignes_excel(chemin: str) -> list[dict]:
    classeur = load_workbook(chemin, read_only=True, data_only=True)
    feuille = classeur.active
    lignes_brutes = list(feuille.iter_rows(values_only=True))
    entetes = [str(c).strip() for c in lignes_brutes[0]]
    lignes = []
    for ligne in lignes_brutes[1:]:
        if not any(ligne):
            continue
        d = dict(zip(entetes, ligne))
        lignes.append({
            "universite": (d.get("Université") or "").strip(),
            "ville": (d.get("Ville") or "").strip(),
            "composante": (d.get("Composante") or "").strip(),
            "domaine": (d.get("Domaine") or "").strip(),
            "mention": (d.get("Mention") or "").strip(),
            "parcours": (d.get("Parcours/Filière") or "").strip(),
        })
    return lignes


def importer(chemin_excel: str, dry_run: bool = False) -> Rapport:
    rapport = Rapport()
    lignes = lire_lignes_excel(chemin_excel)

    with Session(engine) as session:
        # --- Index en memoire de l'existant, cle = nom normalise ---
        universites_par_nom = {normaliser(u.nom): u for u in session.exec(select(Universite)).all()}
        domaines_par_nom = {normaliser(d.nom): d for d in session.exec(select(Domaine)).all()}
        mentions_par_nom = {normaliser(m.nom): m for m in session.exec(select(Mention)).all()}
        facultes_par_cle = {
            (f.universite_id, normaliser(f.nom)): f for f in session.exec(select(Faculte)).all()
        }
        filieres_par_faculte = defaultdict(dict)  # faculte_id -> {nom_normalise: Filiere}
        for fil in session.exec(select(Filiere)).all():
            filieres_par_faculte[fil.faculte_id][normaliser(fil.nom)] = fil
        programmes_existants = {
            (p.universite_id, p.filiere_id) for p in session.exec(select(ProgrammeUniversitaire)).all()
        }

        # --- Universites deja "curatees" = ont au moins une Filiere
        #     existante rattachee (peu importe la faculte). Determine
        #     dynamiquement plutot que code en dur : reste correct si
        #     Fianarantsoa/Mahajanga/Toliara/Antsiranana recoivent un
        #     jour des Filiere par un autre moyen. ---
        universites_curatees_ids = {
            fac.universite_id
            for fac in session.exec(select(Faculte)).all()
            if filieres_par_faculte.get(fac.id)
        }

        # --- Ne garder que les lignes du perimetre (6 universites
        #     publiques), et filtrer les lignes dont l'universite du
        #     fichier ne correspond a AUCUNE Universite en base
        #     (ne devrait pas arriver pour le perimetre public, mais
        #     on ne veut jamais planter silencieusement / inventer). ---
        lignes_retenues = []
        for ligne in lignes:
            cle = normaliser(ligne["universite"])
            if cle not in PERIMETRE_UNIVERSITES_PUBLIQUES:
                rapport.universites_hors_perimetre.add(ligne["universite"])
                continue
            lignes_retenues.append(ligne)

        # === ETAPE 1 : Domaine (national, dedup par nom normalise) ===
        for ligne in lignes_retenues:
            cle = normaliser(ligne["domaine"])
            if not cle or cle in domaines_par_nom:
                continue
            dom = Domaine(nom=ligne["domaine"])
            if not dry_run:
                session.add(dom)
                session.commit()
                session.refresh(dom)
            domaines_par_nom[cle] = dom
            rapport.domaines_crees.append(ligne["domaine"])

        # === ETAPE 2 : Mention (nationale, dedup par nom normalise) +
        #     rattachement Domaine SEULEMENT si non ambigu ===
        domaine_textes_par_mention = defaultdict(set)
        for ligne in lignes_retenues:
            cle_mention = normaliser(ligne["mention"])
            if not cle_mention:
                continue
            domaine_textes_par_mention[cle_mention].add(ligne["domaine"])

        for ligne in lignes_retenues:
            cle_mention = normaliser(ligne["mention"])
            if not cle_mention:
                continue
            if cle_mention not in mentions_par_nom:
                ment = Mention(nom=ligne["mention"])
                if not dry_run:
                    session.add(ment)
                    session.commit()
                    session.refresh(ment)
                mentions_par_nom[cle_mention] = ment
                rapport.mentions_creees.append(ligne["mention"])

        for cle_mention, textes_domaine in domaine_textes_par_mention.items():
            mention = mentions_par_nom[cle_mention]
            if mention.domaine_id is not None:
                continue  # deja rattachee (import precedent ou admin) : ne jamais ecraser
            textes_normalises = {normaliser(t) for t in textes_domaine if t}
            if len(textes_normalises) != 1:
                rapport.mentions_domaine_ambigu[mention.nom] = textes_domaine
                continue
            domaine = domaines_par_nom.get(next(iter(textes_normalises)))
            if domaine is None:
                continue
            mention.domaine_id = domaine.id
            if not dry_run:
                session.add(mention)
                session.commit()
            rapport.mentions_domaine_rattache.append((mention.nom, domaine.nom))

        # === ETAPE 3 : selon curatee ou vide ===
        for ligne in lignes_retenues:
            cle_universite = normaliser(ligne["universite"])
            universite = universites_par_nom.get(cle_universite)
            if universite is None:
                # Ne devrait pas arriver (verifie a l'etape perimetre),
                # garde-fou pour ne jamais planter sur une donnee
                # inattendue.
                rapport.lignes_sans_correspondance.append(
                    (ligne["universite"], ligne["composante"], ligne["domaine"], ligne["mention"], ligne["parcours"])
                )
                continue

            mention = mentions_par_nom.get(normaliser(ligne["mention"]))

            if universite.id in universites_curatees_ids:
                # --- Universite curatee a la main : jamais de nouvelle
                #     Filiere. On cherche uniquement une correspondance
                #     EXACTE (nom normalise) parmi les Filiere deja
                #     rattachees a cette universite, pour lui rattacher
                #     un Domaine via sa Mention si elle n'en a pas. ---
                cle_parcours = normaliser(ligne["parcours"])
                filiere_existante = None
                for fac in session.exec(select(Faculte).where(Faculte.universite_id == universite.id)).all():
                    candidate = filieres_par_faculte.get(fac.id, {}).get(cle_parcours)
                    if candidate is not None:
                        filiere_existante = candidate
                        break

                if filiere_existante is None:
                    rapport.lignes_sans_correspondance.append(
                        (ligne["universite"], ligne["composante"], ligne["domaine"], ligne["mention"], ligne["parcours"])
                    )
                    continue

                if filiere_existante.mention_id is None and mention is not None:
                    filiere_existante.mention_id = mention.id
                    if not dry_run:
                        session.add(filiere_existante)
                        session.commit()
                    rapport.filieres_existantes_rattachees_domaine.append(
                        (ligne["universite"], filiere_existante.nom, mention.domaine.nom if mention.domaine else "(aucun)")
                    )
                continue

            # --- Universite "vide" : import complet ---
            cle_faculte = (universite.id, normaliser(ligne["composante"]))
            faculte = facultes_par_cle.get(cle_faculte)
            if faculte is None:
                faculte = Faculte(nom=ligne["composante"], universite_id=universite.id)
                if not dry_run:
                    session.add(faculte)
                    session.commit()
                    session.refresh(faculte)
                facultes_par_cle[cle_faculte] = faculte
                rapport.facultes_creees.append((ligne["universite"], ligne["composante"]))

            cle_filiere = normaliser(ligne["parcours"])
            filiere = filieres_par_faculte[faculte.id].get(cle_filiere) if faculte.id else None
            if filiere is None:
                filiere = Filiere(
                    nom=ligne["parcours"],
                    faculte_id=faculte.id,
                    mention_id=mention.id if mention else None,
                )
                if not dry_run:
                    session.add(filiere)
                    session.commit()
                    session.refresh(filiere)
                filieres_par_faculte[faculte.id][cle_filiere] = filiere
                rapport.filieres_creees.append((ligne["universite"], ligne["composante"], ligne["parcours"]))

            cle_programme = (universite.id, filiere.id)
            if cle_programme not in programmes_existants:
                if not dry_run:
                    session.add(ProgrammeUniversitaire(universite_id=universite.id, filiere_id=filiere.id))
                    session.commit()
                programmes_existants.add(cle_programme)
                rapport.programmes_crees += 1

        if dry_run:
            session.rollback()

    return rapport


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--fichier", default=FICHIER_PAR_DEFAUT)
    parser.add_argument("--dry-run", action="store_true", help="N'ecrit rien en base, affiche seulement ce qui serait fait.")
    args = parser.parse_args()

    rapport = importer(args.fichier, dry_run=args.dry_run)
    rapport.imprimer()
